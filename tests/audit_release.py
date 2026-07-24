from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "results"
SUPPLEMENT = ROOT / "reference" / "supplementary_data"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def main() -> None:
    failures: list[str] = []
    checks: dict[str, object] = {}

    required = [
        ROOT / "README.md",
        ROOT / "LICENSE",
        ROOT / "data" / "README.md",
        ROOT / "data" / "derived" / "dynamic_program_candidates.csv",
        ROOT / "data" / "derived" / "clean_D1D2_frozen_genes.csv",
        ROOT / "data" / "manifests" / "input_objects.csv",
        RESULTS / "clean_programs" / "clean_D1D2_gene_level_audit.csv",
        RESULTS / "bulk_recurrence" / "bulk_ROC_AP_permutation_statistics.csv",
        SUPPLEMENT / "Supplementary_Methods.docx",
        SUPPLEMENT / "Supplementary_Data_Index.csv",
        SUPPLEMENT / "Supplementary_Data_S4_Trajectory_Dynamics.xlsx",
    ]
    missing = [str(path.relative_to(ROOT)) for path in required if not path.exists()]
    checks["required_files_missing"] = missing
    failures.extend(f"missing required file: {path}" for path in missing)

    clean_rows = read_csv(ROOT / "data" / "derived" / "clean_D1D2_frozen_genes.csv")
    counts: dict[str, int] = {}
    for row in clean_rows:
        counts[row["clean_module"]] = counts.get(row["clean_module"], 0) + 1
    checks["frozen_gene_counts"] = counts
    if counts != {"clean_D1": 42, "clean_D2": 975}:
        failures.append(f"unexpected frozen gene counts: {counts}")

    dynamic_rows = read_csv(ROOT / "data" / "derived" / "dynamic_program_candidates.csv")
    checks["dynamic_candidate_count"] = len(dynamic_rows)
    if len(dynamic_rows) != 1118:
        failures.append(f"expected 1118 dynamic candidates, found {len(dynamic_rows)}")

    index_rows = read_csv(SUPPLEMENT / "Supplementary_Data_Index.csv")
    index_status: dict[str, dict[str, object]] = {}
    for row in index_rows:
        path = SUPPLEMENT / row["file"]
        exists = path.exists()
        observed_hash = sha256(path) if exists else ""
        hash_match = exists and observed_hash == row["sha256"]
        worksheet_match = True
        observed_worksheets = 0
        if exists and path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=False)
            observed_worksheets = len(workbook.sheetnames)
            workbook.close()
            worksheet_match = observed_worksheets == int(row["worksheet_count"])
        index_status[row["file"]] = {
            "exists": exists,
            "sha256_match": hash_match,
            "worksheet_count": observed_worksheets,
            "worksheet_count_match": worksheet_match,
        }
        if not exists:
            failures.append(f"supplementary index target missing: {row['file']}")
        elif not hash_match:
            failures.append(f"supplementary index checksum mismatch: {row['file']}")
        if not worksheet_match:
            failures.append(f"supplementary worksheet-count mismatch: {row['file']}")
    checks["supplementary_index_status"] = index_status

    donor_summary = read_csv(RESULTS / "discovery_donor_level_summary.csv")
    donor_tests = read_csv(RESULTS / "discovery_donor_level_tests.csv")
    doublet_summary = read_csv(
        RESULTS / "QC_Harmony" / "doublet_sensitivity_donor_level_summary.csv"
    )
    s4 = load_workbook(
        SUPPLEMENT / "Supplementary_Data_S4_Trajectory_Dynamics.xlsx",
        read_only=True,
        data_only=True,
    )
    summary_rows = list(
        s4["Donor_Trajectory_Summary"].iter_rows(min_row=2, values_only=True)
    )
    effect_rows = list(s4["Donor_Group_Effects"].iter_rows(min_row=2, values_only=True))
    s4.close()

    csv_d2 = {row["sample"]: float(row["median_clean_D2"]) for row in donor_summary}
    s4_d2 = {str(row[0]): float(row[5]) for row in summary_rows}
    doublet_d2 = {
        row["sample"]: float(row["median_D2"])
        for row in doublet_summary
        if row["version"] == "all_current_cells"
    }
    donor_d2_match = (
        csv_d2.keys() == s4_d2.keys() == doublet_d2.keys()
        and all(abs(csv_d2[key] - s4_d2[key]) < 1e-12 for key in csv_d2)
        and all(abs(csv_d2[key] - doublet_d2[key]) < 1e-12 for key in csv_d2)
    )

    csv_effect = next(row for row in donor_tests if row["metric"] == "median_clean_D2")
    s4_effect = next(row for row in effect_rows if row[0] == "median_clean_D2")
    effect_fields = [
        "normal_median",
        "keloid_median",
        "keloid_minus_normal_mean",
        "wilcoxon_exact_p",
        "exact_permutation_p",
        "cliffs_delta",
        "keloid_higher_pair_fraction",
        "tied_pair_fraction",
        "cross_group_pair_count",
    ]
    effect_match = all(
        abs(float(csv_effect[field]) - float(s4_effect[index + 1])) < 1e-12
        for index, field in enumerate(effect_fields)
    )
    checks["clean_D2_cross_file_consistency"] = {
        "donor_summary_match": donor_d2_match,
        "donor_group_effect_match": effect_match,
        "donor_count": len(csv_d2),
    }
    if not donor_d2_match:
        failures.append("clean D2 donor summaries differ across S4 and result CSV files")
    if not effect_match:
        failures.append("clean D2 donor group effects differ between S4 and result CSV")

    stale_patterns = {
        "absolute_windows_path": re.compile(r"(?<![A-Za-z])[A-Za-z]:[\\/]"),
        "legacy_package_path": re.compile(r"06_Reanalysis_Audit|04_Reproducible_Scripts|final_inputs"),
        "legacy_story_terms": re.compile(r"KFPS|RTI-M|high-confidence|M1-M5"),
    }
    source_files = list((ROOT / "analysis").glob("*")) + list((ROOT / "tests").glob("*.py"))
    source_hits: dict[str, list[str]] = {name: [] for name in stale_patterns}
    for path in source_files:
        if path.suffix.lower() not in {".r", ".py", ".ps1", ".sh"}:
            continue
        if path.resolve() == Path(__file__).resolve():
            continue
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        for name, pattern in stale_patterns.items():
            if pattern.search(text):
                source_hits[name].append(path.name)
    checks["source_hygiene"] = source_hits
    for name, hits in source_hits.items():
        if hits:
            failures.append(f"{name} found in: {', '.join(sorted(hits))}")

    manifest = read_csv(ROOT / "data" / "manifests" / "input_objects.csv")
    env_paths = {
        "GSE191067_standard_reclustered_scored.rds": os.getenv("KFT_GSE191067_OBJECT", ""),
        "Step04B_clean_gene_symbol_log2TPM_plus1_matrix.csv": os.getenv("PRJNA813172_LOG2TPM", ""),
        "Step04B_clean_metadata_matched_to_expression.csv": os.getenv("PRJNA813172_METADATA", ""),
        "filereport_read_run_PRJNA813172.tsv": str(
            ROOT / "data" / "manifests" / "filereport_read_run_PRJNA813172.tsv"
        ),
    }
    input_root = Path(os.getenv("KFT_INPUT_ROOT", ROOT / "data" / "frozen_inputs"))
    verified_inputs: dict[str, str] = {}
    for row in manifest:
        name = row["file"]
        candidate = Path(env_paths.get(name, "")) if env_paths.get(name) else input_root / name
        if candidate.exists():
            observed = sha256(candidate)
            verified_inputs[name] = "match" if observed == row["sha256"] else "mismatch"
            if observed != row["sha256"]:
                failures.append(f"checksum mismatch: {name}")
        else:
            verified_inputs[name] = "not supplied; manifest only"
    checks["input_checksum_status"] = verified_inputs

    report = {"status": "PASS" if not failures else "FAIL", "checks": checks, "failures": failures}
    report_path = Path(os.getenv("KFT_RELEASE_AUDIT_REPORT", ROOT / "tests" / "release_audit_report.json"))
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    if failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
